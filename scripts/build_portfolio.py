#!/usr/bin/env python3
"""Build INT8SystolicGEMMA_Portfolio.xlsx"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

wb = Workbook()

# ── Palette ───────────────────────────────────────────────────────────────────
HDR_FILL    = PatternFill("solid", fgColor="1F3864")
HDR_FONT    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
SUBHDR_FILL = PatternFill("solid", fgColor="2E75B6")
SUBHDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
ALT_FILL    = PatternFill("solid", fgColor="EEF3FA")
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
PASS_FILL   = PatternFill("solid", fgColor="C6EFCE")
PASS_FONT   = Font(name="Arial", bold=True, color="276221", size=10)
ZERO_FILL   = PatternFill("solid", fgColor="FFF2CC")
NEG_FONT    = Font(name="Arial", color="FF0000", size=9)
NORM_FONT   = Font(name="Arial", size=10)
BOLD_FONT   = Font(name="Arial", bold=True, size=10)
TITLE_FONT  = Font(name="Arial", bold=True, size=16, color="1F3864")
SUB_FONT    = Font(name="Arial", italic=True, size=11, color="2E75B6")


def tb():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


def header_row(ws, row, values, col_start=1):
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=col_start + i, value=v)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = tb()


def data_cell(ws, row, col, val, alt=False, font=None, halign="center"):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = ALT_FILL if alt else WHITE_FILL
    c.font = font or NORM_FONT
    c.alignment = Alignment(horizontal=halign, vertical="center", indent=(1 if halign == "left" else 0))
    c.border = tb()
    return c


def section_banner(ws, row, text, n_cols, color="1F3864"):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 18


def write_matrix(ws, top_row, left_col, matrix, label, show_zero_fill=False, label_color="2E75B6"):
    n = len(matrix)
    lf = PatternFill("solid", fgColor=label_color)
    ws.merge_cells(start_row=top_row, start_column=left_col,
                   end_row=top_row, end_column=left_col + n)
    lc = ws.cell(row=top_row, column=left_col, value=label)
    lc.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    lc.fill = lf
    lc.alignment = Alignment(horizontal="left", vertical="center")
    lc.border = tb()

    ch = top_row + 1
    ws.cell(row=ch, column=left_col, value="").fill = HDR_FILL
    ws.cell(row=ch, column=left_col).border = tb()
    for ci in range(n):
        c = ws.cell(row=ch, column=left_col + 1 + ci, value=f"c{ci}")
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center"); c.border = tb()

    for ri in range(n):
        rl = ws.cell(row=ch + 1 + ri, column=left_col, value=f"r{ri}")
        rl.fill = SUBHDR_FILL; rl.font = SUBHDR_FONT
        rl.alignment = Alignment(horizontal="center", vertical="center"); rl.border = tb()
        for ci in range(n):
            val = matrix[ri][ci]
            c = ws.cell(row=ch + 1 + ri, column=left_col + 1 + ci, value=val)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = tb()
            if show_zero_fill and val == 0:
                c.fill = ZERO_FILL
                c.font = Font(name="Arial", size=9, color="AAAAAA", italic=True)
            elif val < 0:
                c.fill = ALT_FILL if ri % 2 else WHITE_FILL
                c.font = NEG_FONT
            else:
                c.fill = ALT_FILL if ri % 2 else WHITE_FILL
                c.font = Font(name="Arial", size=9)

    return ch + 1 + n + 1  # next free row


def write_bias_row(ws, row, left_col, bias, n):
    ws.cell(row=row, column=left_col, value="Bias (per col)").fill = SUBHDR_FILL
    ws.cell(row=row, column=left_col).font = SUBHDR_FONT
    ws.cell(row=row, column=left_col).alignment = Alignment(horizontal="left");
    ws.cell(row=row, column=left_col).border = tb()
    for i, v in enumerate(bias):
        c = ws.cell(row=row, column=left_col + 1 + i, value=v)
        c.fill = ALT_FILL; c.font = NORM_FONT
        c.alignment = Alignment(horizontal="center"); c.border = tb()


def set_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w


# ═════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Project Overview
# ═════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Project Overview"
ws1.sheet_properties.tabColor = "1F3864"
ws1.sheet_view.showGridLines = False

r = 1
ws1.merge_cells(f"A{r}:H{r}")
ws1[f"A{r}"].value = "INT8 Systolic GEMM Accelerator"
ws1[f"A{r}"].font = TITLE_FONT
ws1[f"A{r}"].alignment = Alignment(horizontal="left", vertical="center")
ws1.row_dimensions[r].height = 32

r = 2
ws1.merge_cells(f"A{r}:H{r}")
ws1[f"A{r}"].value = "FPGA Implementation of  Y = ReLU( A × B + Bias )  for Neural Network Inference"
ws1[f"A{r}"].font = SUB_FONT
ws1[f"A{r}"].alignment = Alignment(horizontal="left", vertical="center")
ws1.row_dimensions[r].height = 22

r = 4
section_banner(ws1, r, "Design Parameters", 8)
r += 1
header_row(ws1, r, ["Parameter", "Default", "Description"])
ws1.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
r += 1
for i, (p, d, desc) in enumerate([
    ("N",                "2 – 16", "Array dimension; instantiates N×N processing elements"),
    ("DATA_WIDTH",       "8",      "Input precision — signed INT8"),
    ("ACC_WIDTH",        "32",     "Accumulator width — signed INT32"),
    ("PIPELINE_PRODUCT", "0 / 1",  "Register product before accumulate (+1 cycle latency per PE)"),
    ("PIPELINE_DSP",     "0 / 1",  "2-stage DSP pipeline matching Cyclone V DSP blocks (+2 cycle latency)"),
    ("ENABLE_BIAS",      "0 / 1",  "Add per-column bias value to each accumulator output"),
    ("ENABLE_RELU",      "0 / 1",  "Clamp negative post-bias values to zero (ReLU activation)"),
]):
    alt = i % 2 == 1
    data_cell(ws1, r, 1, p, alt, BOLD_FONT, "left")
    data_cell(ws1, r, 2, d, alt)
    data_cell(ws1, r, 3, desc, alt, halign="left")
    ws1.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
    r += 1

r += 1
section_banner(ws1, r, "Module Hierarchy", 8)
r += 1
header_row(ws1, r, ["Module", "File", "Role"])
ws1.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
r += 1
for i, (m, f, role) in enumerate([
    ("matrix_accelerator", "rtl/matrix_accelerator.sv",
     "Top-level integrator: BRAM banks, FSM, systolic array, 2-stage post-process pipeline"),
    ("systolic_array",     "rtl/systolic_array.sv",
     "N×N grid of PEs — A data flows left→right, B data flows top→bottom"),
    ("pe",                 "rtl/pe.sv",
     "INT8×INT8 multiply → sign-extended → INT32 accumulate; 3 configurable pipeline modes"),
    ("controller_fsm",     "rtl/controller_fsm.sv",
     "7-state FSM; diagonal wave-front scheduling to feed correct (A,B) pairs each cycle"),
    ("post_process",       "rtl/post_process.sv",
     "Combinational: biased = acc + bias; y = (biased < 0 && ENABLE_RELU) ? 0 : biased"),
    ("bram_model",         "rtl/bram_model.sv",
     "Dual-port BRAM abstraction: behavioral model in simulation, altsyncram M10K in synthesis"),
]):
    alt = i % 2 == 1
    data_cell(ws1, r, 1, m, alt, BOLD_FONT, "left")
    data_cell(ws1, r, 2, f, alt, halign="left")
    data_cell(ws1, r, 3, role, alt, halign="left")
    ws1.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
    r += 1

r += 1
section_banner(ws1, r, "Controller FSM States", 8)
r += 1
header_row(ws1, r, ["State", "Cycles", "Purpose"])
ws1.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
r += 1
for i, (st, cy, pur) in enumerate([
    ("IDLE",          "—",          "Wait for start pulse"),
    ("CLEAR",         "1",          "Assert clear to all PEs — zero all accumulators and pipeline registers"),
    ("RUN",           "T_MAX + 1",  "Diagonal wave-front: row i reads A[i][t−i], col j reads B[t−j][j]"),
    ("WRITE_OUTPUT",  "N²",         "Sequential scan of acc_out[row][col]; feeds post-process pipeline"),
    ("WRITE_FLUSH1",  "1",          "Drain post-process stage 1 (bias add pipeline register)"),
    ("WRITE_FLUSH2",  "1",          "Drain post-process stage 2; last result written to BRAM_C"),
    ("DONE",          "1",          "Assert done=1 for one cycle; auto-return to IDLE"),
]):
    alt = i % 2 == 1
    data_cell(ws1, r, 1, st, alt, BOLD_FONT)
    data_cell(ws1, r, 2, cy, alt)
    data_cell(ws1, r, 3, pur, alt, halign="left")
    ws1.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
    r += 1

r += 1
section_banner(ws1, r, "Cycle Count Formula", 8)
r += 1
for i, (term, formula, note) in enumerate([
    ("Busy Cycles",         "1  +  (T_MAX + 1)  +  N²  +  2",
     "CLEAR + RUN + WRITE_OUTPUT + FLUSH1 + FLUSH2"),
    ("T_MAX",               "3N − 3  +  PE_pipeline_latency",
     "Total time steps for all wave-front products to propagate"),
    ("PE_pipeline_latency", "2 if PIPELINE_DSP=1,  else PIPELINE_PRODUCT (0 or 1)",
     "Extra cycles before PE accumulates; matches DSP input+output register depth"),
]):
    alt = i % 2 == 1
    data_cell(ws1, r, 1, term, alt, BOLD_FONT, "left")
    data_cell(ws1, r, 2, formula, alt, halign="left")
    ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    data_cell(ws1, r, 6, note, alt, Font(name="Arial", size=9, italic=True, color="555555"), "left")
    ws1.merge_cells(start_row=r, start_column=6, end_row=r, end_column=8)
    r += 1

r += 1
section_banner(ws1, r, "FPGA Target & Toolchain", 8)
r += 1
header_row(ws1, r, ["Item", "Value"])
ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
r += 1
for i, (item, val) in enumerate([
    ("Device",           "Intel Cyclone V — 5CGXFC9E7F35C8"),
    ("Synthesis Tool",   "Quartus Prime 20.1.1 Lite Edition"),
    ("Simulator",        "ModelSim Intel FPGA Starter Edition 2020.1"),
    ("Clock Constraint", "100 MHz (10 ns period) — constraints/clock.sdc"),
    ("Block Memory",     "Altera M10K (altsyncram dual-port) — one bank per row of A, one per col of B, one flat bank for output C"),
    ("Virtual Pins",     "All top-level I/O marked VIRTUAL_PIN — allows synthesis without physical pin assignment"),
    ("Language",         "SystemVerilog (IEEE 1800)"),
    ("Test Vectors",     "Generated by scripts/gen_vectors.py — golden reference: scripts/golden_model.py"),
]):
    alt = i % 2 == 1
    data_cell(ws1, r, 1, item, alt, BOLD_FONT, "left")
    data_cell(ws1, r, 2, val, alt, halign="left")
    ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    ws1.row_dimensions[r].height = 18
    r += 1

set_widths(ws1, [22, 32, 50, 12, 12, 12, 12, 12])

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Simulation Results
# ═════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Simulation Results")
ws2.sheet_properties.tabColor = "375623"
ws2.sheet_view.showGridLines = False

r = 1
ws2.merge_cells("A1:L1")
ws2["A1"].value = "Simulation Regression Results — ModelSim Intel FPGA Starter 2020.1"
ws2["A1"].font = TITLE_FONT
ws2["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws2.row_dimensions[1].height = 28
r = 2
ws2.merge_cells(f"A{r}:L{r}")
ws2[f"A{r}"].value = ("All 14 simulation runs executed against Python golden model (scripts/golden_model.py) "
                       "— PASS = zero output mismatches, zero ModelSim errors/warnings")
ws2[f"A{r}"].font = SUB_FONT
ws2[f"A{r}"].alignment = Alignment(horizontal="left", vertical="center")
ws2.row_dimensions[r].height = 18
r = 4

header_row(ws2, r, ["#","Test Name","N","Pipeline Mode","ENABLE_BIAS","ENABLE_RELU",
                     "T_MAX","Busy Cycles","Outputs Checked","Errors","Warnings","Result"])
ws2.row_dimensions[r].height = 20
r += 1
sim_rows = [
    (1,  "signed_basic",            2,  "Baseline",     "No","No",  3,  11,  4,   0, 0),
    (2,  "identity",                2,  "Baseline",     "No","No",  3,  11,  4,   0, 0),
    (3,  "zero",                    2,  "Baseline",     "No","No",  3,  11,  4,   0, 0),
    (4,  "signed_basic",            4,  "Baseline",     "No","No",  9,  29,  16,  0, 0),
    (5,  "identity",                4,  "Baseline",     "No","No",  9,  29,  16,  0, 0),
    (6,  "zero",                    4,  "Baseline",     "No","No",  9,  29,  16,  0, 0),
    (7,  "signed_random",           4,  "Baseline",     "No","No",  9,  29,  16,  0, 0),
    (8,  "int8_minmax_stress",      4,  "Baseline",     "No","No",  9,  29,  16,  0, 0),
    (9,  "bias_zero",               2,  "DSP Pipeline", "Yes","Yes",5,  13,  4,   0, 0),
    (10, "bias_positive",           4,  "DSP Pipeline", "Yes","Yes",11, 31,  16,  0, 0),
    (11, "bias_negative",           4,  "DSP Pipeline", "Yes","Yes",11, 31,  16,  0, 0),
    (12, "relu_basic",              8,  "DSP Pipeline", "Yes","Yes",21, 91,  64,  0, 0),
    (13, "signed_random_bias_relu", 8,  "DSP Pipeline", "Yes","Yes",21, 91,  64,  0, 0),
    (14, "signed_random_bias_relu", 16, "DSP Pipeline", "Yes","Yes",45, 307, 256, 0, 0),
]
for i, row_data in enumerate(sim_rows):
    alt = i % 2 == 1
    for ci, val in enumerate(row_data, 1):
        data_cell(ws2, r, ci, val, alt)
    pc = ws2.cell(row=r, column=12, value="PASS")
    pc.fill = PASS_FILL; pc.font = PASS_FONT
    pc.alignment = Alignment(horizontal="center", vertical="center"); pc.border = tb()
    r += 1

r += 1
for label, val, color in [("Total Tests:", 14, "000000"), ("Passed:", 14, "276221"), ("Failed:", 0, "000000")]:
    ws2.cell(row=r, column=1, value=label).font = Font(name="Arial", bold=True, size=10, color=color)
    ws2.cell(row=r, column=2, value=val).font = Font(name="Arial", bold=True, size=10, color=color)
    ws2.cell(row=r, column=2).alignment = Alignment(horizontal="center")
    r += 1

r += 1
section_banner(ws2, r, "Test Case Descriptions", 12)
r += 1
header_row(ws2, r, ["Test Name", "Purpose / What It Verifies"])
ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=12)
r += 1
for i, (name, purpose) in enumerate([
    ("signed_basic",            "Hand-crafted signed integers; verifies basic 2×2 and 4×4 signed INT8 MAC correctness"),
    ("identity",                "A × I = A; verifies correct PE diagonal propagation and that no data is dropped or skewed"),
    ("zero",                    "A × 0 = 0; verifies CLEAR state properly zeros all PE accumulators before computation"),
    ("signed_random",           "Deterministic random INT8 values (seed 0x1A57+N); broad arithmetic coverage across full signed range"),
    ("int8_minmax_stress",      "Corner values ±128, ±127, ±64, 0; verifies INT32 accumulator headroom — max 4 products of ±128 = ±65536 ≪ INT32_MAX"),
    ("bias_zero",               "ENABLE_BIAS=1, ENABLE_RELU=1 with bias=0; confirms bias path active but doesn't corrupt with zero-bias inputs"),
    ("bias_positive",           "Positive bias shifts outputs up; some raw negatives become positive after bias, then pass ReLU"),
    ("bias_negative",           "All-ones A&B (raw C=4 per element); negative bias makes all values negative; ReLU clamps all to 0"),
    ("relu_basic",              "A=all rows [−1,0,…]; B=all rows [8,9,…]; all dot products = −8 → bias=0 → ReLU clamps all 64 outputs to 0"),
    ("signed_random_bias_relu", "Full AI inference mode: random INT8 weights, randomized bias (seed 0xB1A5+N), ReLU activation; all N² outputs verified"),
]):
    alt = i % 2 == 1
    data_cell(ws2, r, 1, name, alt, BOLD_FONT, "left")
    c2 = ws2.cell(row=r, column=2, value=purpose)
    c2.fill = ALT_FILL if alt else WHITE_FILL
    c2.font = NORM_FONT
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    c2.border = tb()
    ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=12)
    ws2.row_dimensions[r].height = 28
    r += 1

set_widths(ws2, [5, 26, 5, 18, 12, 12, 8, 12, 16, 8, 10, 10])

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 3 — N=2 Vectors
# ═════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("N=2 Vectors")
ws3.sheet_properties.tabColor = "7030A0"
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:K1")
ws3["A1"].value = "N=2 Test Vectors  (DATA_WIDTH=8, ACC_WIDTH=32, Baseline pipeline unless noted)"
ws3["A1"].font = TITLE_FONT
ws3["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws3.row_dimensions[1].height = 28
r = 3

n2_tests = [
    ("Test 1 — signed_basic",  "Verifies basic 2×2 signed INT8 MAC",
     [[1,2],[3,4]], [[5,6],[7,8]], [0,0], [[19,22],[43,50]], False),
    ("Test 2 — identity",      "A × I = A  (B = identity matrix)",
     [[1,2],[3,4]], [[1,0],[0,1]], [0,0], [[1,2],[3,4]], False),
    ("Test 3 — zero",          "A × 0 = 0  (B = zero matrix)",
     [[1,2],[3,4]], [[0,0],[0,0]], [0,0], [[0,0],[0,0]], False),
    ("Test 9 — bias_zero",     "ENABLE_BIAS=1, ENABLE_RELU=1, bias=[0,0].  Y=ReLU(A×B+0)=A×B (DSP Pipeline)",
     [[1,2],[2,3]], [[1,2],[3,1]], [0,0], [[7,4],[11,7]], True),
]
for label, note, A, B, bias, C, has_post in n2_tests:
    ws3.merge_cells(f"A{r}:K{r}")
    c = ws3.cell(row=r, column=1, value=label)
    c.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="7030A0")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws3.row_dimensions[r].height = 20
    r += 1
    ws3.merge_cells(f"A{r}:K{r}")
    c = ws3.cell(row=r, column=1, value=note)
    c.font = Font(name="Arial", italic=True, size=9, color="555555")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    r += 1

    r_a = write_matrix(ws3, r, 1, A, "Matrix A (INT8)")
    write_matrix(ws3, r, 4, B, "Matrix B (INT8)")
    write_bias_row(ws3, r_a - 1, 1, bias, 2)
    c_lbl = "Y = ReLU(A×B+bias)" if has_post else "C = A×B (INT32)"
    write_matrix(ws3, r, 7, C, c_lbl, show_zero_fill=has_post, label_color="375623")
    r = r_a + 2

set_widths(ws3, [14, 8, 8, 4, 8, 8, 4, 8, 8, 8, 8])

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 4 — N=4 Vectors
# ═════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("N=4 Vectors")
ws4.sheet_properties.tabColor = "ED7D31"
ws4.sheet_view.showGridLines = False

ws4.merge_cells("A1:S1")
ws4["A1"].value = "N=4 Test Vectors  (DATA_WIDTH=8, ACC_WIDTH=32)"
ws4["A1"].font = TITLE_FONT
ws4["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws4.row_dimensions[1].height = 28
r = 3

A4b = [[1,-2,3,-4],[5,6,-7,8],[-1,2,-3,4],[7,-8,9,-10]]
B4b = [[2,0,-1,3],[-4,5,6,-7],[8,-9,10,11],[-12,13,-14,15]]
C4b = [[82,-89,73,-10],[-166,197,-151,16],[-82,89,-73,10],[238,-251,175,26]]
A4r = [[-85,-87,115,42],[-63,33,86,-34],[-103,-1,17,108],[-56,-104,-123,10]]
B4r = [[-41,37,24,26],[124,90,-21,65],[-83,33,-98,-82],[-87,89,-6,11]]
C4r = [[-20502,-3442,-11735,-16833],[2495,451,-10429,-6919],
       [-6708,6272,-4765,-2949],[-1261,-14601,12834,1980]]
A4m = [[-128,-127,-64,-1],[0,1,63,64],[126,127,-128,-127],[-64,-1,0,1]]
B4m = [[127,64,-127,1],[-128,-64,63,-1],[1,0,127,64],[-1,126,-128,-64]]
C4m = [[-63,-190,255,-4033],[-129,8000,-128,-65],
       [-255,-16066,-8001,-65],[-8001,-3906,7937,-127]]
A4bn = [[1,1,1,1]]*4
B4bn = [[1,1,1,1]]*4

n4_tests = [
    ("Test 4 — signed_basic",  "4×4 signed mixed integers; verifies multi-row wave-front correctness (Baseline)",
     A4b, B4b, [0,0,0,0], C4b, False),
    ("Test 5 — identity",      "A × I = A; verifies identity at N=4 (Baseline)",
     A4b, [[1,0,0,0],[0,1,0,0],[0,0,1,0],[0,0,0,1]], [0,0,0,0], A4b, False),
    ("Test 6 — zero",          "A × 0 = 0; all 16 outputs must be zero (Baseline)",
     A4b, [[0,0,0,0]]*4, [0,0,0,0], [[0,0,0,0]]*4, False),
    ("Test 7 — signed_random", "Deterministic random INT8 (seed 0x1A57+4); 16 outputs verified (Baseline)",
     A4r, B4r, [0,0,0,0], C4r, False),
    ("Test 8 — int8_minmax_stress", "INT8 corner values ±128, ±127, ±64, 0 in both operands (Baseline)",
     A4m, B4m, [0,0,0,0], C4m, False),
    ("Test 10 — bias_positive", "ENABLE_BIAS=1, ENABLE_RELU=1; positive bias [10,11,12,13] shifts outputs up (DSP Pipeline)",
     A4b, B4b, [10,11,12,13], [[92,0,85,3],[0,208,0,29],[0,100,0,23],[248,0,187,39]], True),
    ("Test 11 — bias_negative", "All-ones A&B (raw C=4); bias [−12..−15] → all sums negative → ReLU clamps to 0 (DSP Pipeline)",
     A4bn, B4bn, [-12,-13,-14,-15], [[0,0,0,0]]*4, True),
]
for label, note, A, B, bias, C, has_post in n4_tests:
    ws4.merge_cells(f"A{r}:S{r}")
    c = ws4.cell(row=r, column=1, value=label)
    c.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="C55A11")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws4.row_dimensions[r].height = 20
    r += 1
    ws4.merge_cells(f"A{r}:S{r}")
    c = ws4.cell(row=r, column=1, value=note)
    c.font = Font(name="Arial", italic=True, size=9, color="555555")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    r += 1

    r_a = write_matrix(ws4, r, 1, A, "Matrix A (INT8)")
    write_matrix(ws4, r, 7, B, "Matrix B (INT8)")
    write_bias_row(ws4, r_a - 1, 1, bias, 4)
    c_lbl = "Y = ReLU(A×B+bias)" if has_post else "C = A×B (INT32)"
    write_matrix(ws4, r, 13, C, c_lbl, show_zero_fill=has_post, label_color="375623")
    r = r_a + 2

set_widths(ws4, [14, 7, 7, 7, 7, 3, 7, 7, 7, 7, 7, 3, 7, 7, 7, 7, 7, 3, 10])

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 5 — N=8 Vectors
# ═════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("N=8 Vectors")
ws5.sheet_properties.tabColor = "C00000"
ws5.sheet_view.showGridLines = False

ws5.merge_cells("A1:Z1")
ws5["A1"].value = "N=8 Test Vectors  (DSP Pipeline, ENABLE_BIAS=1, ENABLE_RELU=1)"
ws5["A1"].font = TITLE_FONT
ws5["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws5.row_dimensions[1].height = 28
r = 3

A8rl = [[-1,0,0,0,0,0,0,0]]*8
B8rl = [[8,9,10,11,12,13,14,15]]*8
C8rl = [[0]*8]*8
A8rn = [[-89,-108,93,106,103,45,34,34],[-35,12,-55,-111,-54,7,-83,-57],
        [-19,126,-106,-111,64,-22,-51,113],[-12,11,-98,-105,-55,-59,-27,69],
        [-89,-115,-128,50,-122,-58,-17,-16],[-79,-10,120,-87,-22,-9,-120,7],
        [68,-120,-2,96,5,-48,-5,69],[-67,65,-121,86,107,35,66,-80]]
B8rn = [[14,-125,18,-20,-67,-18,-119,-120],[47,-10,-116,-105,-80,53,-122,-80],
        [49,-43,-83,20,-34,-67,-66,-26],[13,119,-51,27,2,-113,1,-91],
        [16,58,53,30,-16,87,70,-7],[-21,-96,3,6,-71,74,31,-100],
        [41,58,0,119,-12,-64,-2,124],[-56,101,-62,2,-55,-21,109,-110]]
bias8rn = [-922,1556,944,1546,843,-99,-1611,-20]
C8rn = [[0,29436,2231,26862,5375,0,28367,2491],
        [0,0,9841,0,8374,19724,0,10407],
        [0,8299,0,0,0,31397,8380,0],
        [0,2542,5571,0,5487,10259,6734,7990],
        [0,21175,15108,8128,27612,0,19460,24944],
        [0,0,0,0,4728,7537,187,461],
        [0,17343,7201,14743,5740,0,11558,0],
        [4547,23312,8591,7077,3829,12296,6227,10875]]

n8_tests = [
    ("Test 12 — relu_basic",
     "A = all rows [−1,0,…,0]; B = all rows [8,9,…,15]. All dot products = −8. Bias=0. ReLU clamps all 64 outputs to 0.",
     A8rl, B8rl, [0]*8, C8rl),
    ("Test 13 — signed_random_bias_relu",
     "Random INT8 (seed 0x1A57+8), bias (seed 0xB1A5+8), ReLU. Amber = ReLU-clamped zero. 64 outputs verified.",
     A8rn, B8rn, bias8rn, C8rn),
]
for label, note, A, B, bias, C in n8_tests:
    ws5.merge_cells(f"A{r}:Z{r}")
    c = ws5.cell(row=r, column=1, value=label)
    c.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="C00000")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws5.row_dimensions[r].height = 20
    r += 1
    ws5.merge_cells(f"A{r}:Z{r}")
    c = ws5.cell(row=r, column=1, value=note)
    c.font = Font(name="Arial", italic=True, size=9, color="555555")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ws5.row_dimensions[r].height = 18
    r += 1

    write_matrix(ws5, r, 1, A, "Matrix A (INT8)")
    write_matrix(ws5, r, 11, B, "Matrix B (INT8)")
    write_bias_row(ws5, r + 10, 1, bias, 8)
    r_a = write_matrix(ws5, r, 21, C, "Y = ReLU(A×B+bias) INT32",
                        show_zero_fill=True, label_color="375623")
    r = r_a + 1

set_widths(ws5, [14] + [7]*8 + [3] + [7]*8 + [3] + [8]*8)

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 6 — N=16 Vector
# ═════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("N=16 Vector")
ws6.sheet_properties.tabColor = "833C00"
ws6.sheet_view.showGridLines = False

ws6.merge_cells("A1:AJ1")
ws6["A1"].value = "N=16 Test Vector — signed_random_bias_relu  (DSP Pipeline, ENABLE_BIAS=1, ENABLE_RELU=1)"
ws6["A1"].font = TITLE_FONT
ws6["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws6.row_dimensions[1].height = 28
r = 2
ws6.merge_cells(f"A{r}:AJ{r}")
ws6[f"A{r}"].value = ("Random INT8 inputs (seed 0x1A57+16), bias (seed 0xB1A5+16), ReLU activation. "
                       "256 outputs verified against Python golden model. "
                       "Busy cycles = 307 at 100 MHz = 3.07 µs per inference. Amber = ReLU-clamped zero.")
ws6[f"A{r}"].font = Font(name="Arial", italic=True, size=9, color="555555")
ws6[f"A{r}"].alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
ws6.row_dimensions[r].height = 22
r += 2

A16 = [[112,-17,4,-32,25,-125,114,45,121,49,114,54,28,107,36,38],
       [8,-65,22,-127,114,-110,125,26,3,-39,-15,78,86,46,46,106],
       [-76,-26,-117,106,-3,77,20,-66,37,95,-11,79,-63,87,-47,51],
       [92,-59,-25,42,-66,88,6,39,-115,-99,40,-127,103,95,72,111],
       [-64,-10,-112,-59,-94,73,-51,-22,48,-121,-127,36,-101,18,-97,59],
       [-75,-120,-106,-7,74,91,1,-97,49,108,-90,46,58,-32,106,-51],
       [107,4,27,-120,-35,-108,-59,-29,46,-71,116,-15,-106,29,5,-87],
       [-80,123,-3,121,109,-104,-122,-90,-125,-49,112,-4,-98,19,-15,-61],
       [-112,-113,124,-38,-48,-9,-22,-62,-35,52,-77,126,-98,25,65,122],
       [-12,-5,-35,54,-34,-95,-41,10,-46,-95,6,50,59,104,-47,-27],
       [-44,118,124,96,-62,-53,118,45,17,-65,8,-4,93,84,51,-65],
       [-56,122,3,105,-126,-105,118,106,50,45,3,-25,73,-12,37,-62],
       [-104,-33,-123,-30,-99,0,35,-6,-32,-122,121,24,28,-56,78,4],
       [-43,-121,-31,-66,102,-118,116,-69,-108,115,25,84,-115,-115,1,15],
       [-87,-6,117,115,15,-64,58,29,-43,-8,-54,43,-98,76,-87,-18],
       [3,-58,-120,-98,106,-60,-18,-41,12,-122,-75,36,23,70,-78,-21]]
B16 = [[-43,33,66,36,63,59,-91,-67,92,-80,-52,28,-8,113,19,-101],
       [16,10,-79,45,107,-56,-88,13,-22,-56,102,-35,-120,81,-79,36],
       [-4,38,96,-24,-94,-53,51,-64,-81,112,125,-76,-8,-62,-91,-37],
       [-81,123,112,51,-41,-30,118,76,-115,123,-42,9,81,-66,-106,-41],
       [30,116,121,6,-100,-109,62,-69,-101,-95,7,-31,-97,-72,14,53],
       [-73,44,52,24,-105,53,-54,-117,123,0,-42,33,89,122,100,49],
       [-18,-111,-9,112,124,-99,-18,-78,-62,84,34,94,-11,-66,55,-10],
       [-98,-4,-9,109,54,-31,64,98,-14,-48,-101,72,25,-50,24,-70],
       [-49,-93,-69,-48,-65,39,-60,1,84,-110,-127,-1,-35,122,-68,33],
       [-115,-12,99,-66,66,28,15,-10,-120,-125,-47,-57,41,-103,-15,72],
       [-58,87,45,29,22,-69,107,78,53,-117,-81,-116,-117,-10,-4,113],
       [-36,59,-25,79,0,65,77,-17,73,117,-103,-77,10,-74,-80,-87],
       [-90,-68,12,-48,68,-36,-121,-28,-90,100,-112,-39,61,-110,-94,107],
       [94,102,-80,80,-124,-72,11,-40,79,80,82,41,15,121,7,-62],
       [-66,-119,-116,64,-56,80,77,-1,87,-44,-83,-42,-97,37,-25,-43],
       [93,102,10,-128,125,52,-119,82,110,59,-4,-117,-43,100,67,14]]
bias16 = [1891,1510,-1891,340,1238,101,1595,-970,-1143,219,-1545,1974,212,1489,-1491,105]
C16 = [[0,0,0,19888,23849,0,4325,2656,14806,0,0,0,0,7386,0,0],
       [23164,0,0,2661,22699,0,0,0,2075,20420,0,0,0,0,5742,0],
       [3771,31494,2470,0,0,9461,11415,1897,10138,13144,0,2136,23241,12454,6610,5407],
       [5858,16876,0,6615,8447,1665,0,3003,34164,30198,0,11979,15268,38120,27808,0],
       [44939,170,0,0,0,31672,0,3422,54001,19805,12778,28115,23226,56949,29357,0],
       [0,0,0,0,0,28068,5649,0,0,0,0,4520,22650,0,4800,20869],
       [21272,0,0,8554,0,1621,12906,2604,34674,0,10755,0,0,34367,0,0],
       [30719,62725,13413,9160,0,0,57598,26886,0,0,43318,0,0,0,0,9442],
       [27061,10737,0,0,0,30110,24472,0,17145,48048,17808,0,9617,0,1546,0],
       [19854,21088,0,15829,0,0,11540,16295,0,40265,3874,7755,9498,0,0,0],
       [0,0,0,39693,4035,0,12527,0,0,51261,18665,10905,0,0,0,0],
       [0,0,0,29173,46574,0,8349,27893,0,14565,0,17341,1574,0,0,0],
       [901,0,0,12854,19533,8763,13409,26598,31740,8099,0,0,0,894,10804,14940],
       [12084,0,25961,0,27469,0,41269,0,0,0,4770,0,0,0,21204,2464],
       [21632,34407,15983,20149,0,0,42905,4213,0,52944,44281,13004,15243,0,0,0],
       [45355,4896,0,0,0,0,0,0,13668,5129,4589,22992,0,12394,15271,0]]

write_matrix(ws6, r, 1, A16, "Matrix A (16×16, INT8)")
write_matrix(ws6, r, 19, B16, "Matrix B (16×16, INT8)")
write_bias_row(ws6, r + 18, 1, bias16, 16)
write_matrix(ws6, r, 37, C16, "Y = ReLU(A×B+bias)  16×16 INT32  (amber = ReLU zero)",
             show_zero_fill=True, label_color="375623")

set_widths(ws6, [14] + [6]*16 + [3] + [6]*16 + [3] + [7]*16)

# ── Save ──────────────────────────────────────────────────────────────────────
out = Path(__file__).resolve().parents[1] / "INT8SystolicGEMMA_Portfolio.xlsx"
wb.save(out)
print(f"Saved: {out}")
